"""
Comparativo de BD — conteo de valores no-nulos por columna.

Lee BaseCostosPOSE.xlsx (A2) y los CSV de output/b52
(B52) y genera docs/comparativa.md con tabla de conteos por campo.
La columna A2<->B52 actúa como QA gate antes de cargar a Hetzner.

Uso:
  python src/automatizacion/comparativo_bd.py
"""

import os
import sys
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd

ARCHIVO_BD = "fuentes/compensaciones/BaseCostosPOSE.xlsx"
RUTA_B52 = "output/b52"
SNAPSHOT_PATH = "config/snapshot_bd_conteos.json"
COMPARATIVA_MD = "docs/comparativa.md"


def contar_columnas(ruta: str) -> dict[str, Any]:
    """Cuenta filas totales y no-nulos por columna."""
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    primera_fila = next(ws.iter_rows(min_row=1, max_row=1), None)
    if primera_fila is None:
        wb.close()
        return {"filas": 0, "columnas": {}, "fecha": ""}

    headers = [(c.value or f"COL_{i}") for i, c in enumerate(primera_fila)]
    conteos: dict[str, int] = {h: 0 for h in headers}
    filas = 0

    for row in ws.iter_rows(min_row=2):
        filas += 1
        for i, cell in enumerate(row):
            if i < len(headers) and cell.value is not None:
                v = str(cell.value).strip()
                if v:
                    conteos[headers[i]] += 1

    wb.close()
    return {
        "filas": filas,
        "columnas": conteos,
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def contar_columnas_b52(ruta_dir: str) -> dict[str, Any]:
    """Cuenta filas y no-nulos leyendo todos los CSV de B52."""
    archivos = sorted(
        [
            f
            for f in os.listdir(ruta_dir)
            if f.startswith("BaseCostosPOSE_B52_") and f.endswith(".csv")
        ]
    )
    if not archivos:
        return {"filas": 0, "columnas": {}, "fecha": "sin datos"}

    dfs = [pd.read_csv(os.path.join(ruta_dir, f)) for f in archivos]
    df_total = pd.concat(dfs, ignore_index=True)
    filas = len(df_total)

    conteos: dict[str, int] = {}
    for col in df_total.columns:
        if col.startswith("_"):
            continue
        serie = df_total[col]
        if serie.dtype == object:
            mask = serie.notna() & (serie.str.strip() != "")
        else:
            mask = serie.notna()
        conteos[col] = int(mask.sum())

    return {
        "filas": filas,
        "columnas": conteos,
        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def leer_snapshot() -> dict[str, Any]:
    import json

    if os.path.exists(SNAPSHOT_PATH):
        with open(SNAPSHOT_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def guardar_snapshot(datos: dict[str, Any]) -> None:
    import json

    os.makedirs(os.path.dirname(SNAPSHOT_PATH), exist_ok=True)
    with open(SNAPSHOT_PATH, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)


def _delta_str(d: int) -> str:
    """Formatea el delta entre dos conteos."""
    if d > 0:
        return f"**+{d:,}**"
    if d < 0:
        return f"**{d:,}**"
    return "="


def _estado_b52(v_a2: int, v_b52: int) -> str:
    """Compara conteo A2 vs B52 para columna."""
    diff = v_b52 - v_a2
    if diff == 0:
        return "✅"
    if diff > 0:
        return f"⚠️ +{diff:,}"
    return f"⚠️ {diff:,}"


def generar_md(
    actual: dict[str, Any],
    anterior: dict[str, Any],
    b52: dict[str, Any],
) -> None:
    """Escribe comparativa.md con tabla de conteos por columna."""
    ahora = actual["fecha"]
    headers = list(actual["columnas"].keys())

    filas_b52 = b52.get("filas", 0)
    cols_b52: dict[str, int] = b52.get("columnas", {})
    fecha_b52 = b52.get("fecha", "—")

    lineas: list[str] = [
        "# Comparativo BD — BaseCostosPOSE",
        "",
        f"**Generado:** {ahora}  ",
        f"**Filas A2 (xlsx):** {actual['filas']:,}  ",
        f"**Filas B52 (CSV incr.):** {filas_b52:,}  ",
        f"**B52 leído:** {fecha_b52}",
        "",
        "---",
        "",
        "## Conteo de valores no-nulos por campo",
        "",
    ]

    cols_act = actual["columnas"]
    fecha_act = actual["fecha"]
    cols_ant: dict[str, int] = anterior.get("columnas", {})
    fecha_ant: str = anterior.get("fecha", "")

    if anterior:
        lineas.append(
            f"| Campo"
            f" | A2 ant. ({fecha_ant})"
            f" | A2 act. ({fecha_act})"
            f" | Δ A2"
            f" | B52 incr."
            f" | A2\u2194B52 |"
        )
        lineas.append("|---|---:|---:|:---:|---:|:---:|")
        for h in headers:
            v_ant = cols_ant.get(h, 0)
            v_act = cols_act.get(h, 0)
            v_b52 = cols_b52.get(h, 0)
            lineas.append(
                f"| {h}"
                f" | {v_ant:,}"
                f" | {v_act:,}"
                f" | {_delta_str(v_act - v_ant)}"
                f" | {v_b52:,}"
                f" | {_estado_b52(v_act, v_b52)} |"
            )
    else:
        lineas.append(
            f"| Campo"
            f" | A2 act. ({fecha_act})"
            f" | B52 incr."
            f" | A2\u2194B52 |"
        )
        lineas.append("|---|---:|---:|:---:|")
        for h in headers:
            v_act = cols_act.get(h, 0)
            v_b52 = cols_b52.get(h, 0)
            lineas.append(
                f"| {h}"
                f" | {v_act:,}"
                f" | {v_b52:,}"
                f" | {_estado_b52(v_act, v_b52)} |"
            )

    lineas.append("")
    lineas.append("---")
    lineas.append("")
    lineas.append(
        "**Nota:** A2 = BaseCostosPOSE.xlsx (reservorio estático). "
        "B52 = CSVs particionados UPSERT-ready. "
        "✅ = conteos coinciden. ⚠️ = divergencia a investigar."
    )
    lineas.append("")

    os.makedirs(os.path.dirname(COMPARATIVA_MD), exist_ok=True)
    with open(COMPARATIVA_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lineas))


def main() -> None:
    if not os.path.exists(ARCHIVO_BD):
        print(f"ERROR: no se encontró {ARCHIVO_BD}")
        sys.exit(1)

    anterior = leer_snapshot()

    print(f"  Leyendo {ARCHIVO_BD} ...")
    actual = contar_columnas(ARCHIVO_BD)
    print(f"  Filas A2: {actual['filas']:,}")

    b52: dict[str, Any]
    if os.path.isdir(RUTA_B52):
        print(f"  Leyendo CSVs B52 desde {RUTA_B52} ...")
        b52 = contar_columnas_b52(RUTA_B52)
        print(f"  Filas B52: {b52['filas']:,}")
    else:
        print(f"  AVISO: {RUTA_B52} no encontrado — B52 sin datos.")
        b52 = {"filas": 0, "columnas": {}, "fecha": "—"}

    generar_md(actual, anterior, b52)
    print(f"  Reporte MD guardado en: {COMPARATIVA_MD}")

    guardar_snapshot(actual)
    print(f"  Snapshot guardado en: {SNAPSHOT_PATH}")


if __name__ == "__main__":
    main()
