"""
Comparativo crudo -> ingesta por archivo.

Para cada xlsx en output_ready_for_pq/{segmento}/ busca el archivo
homonimo en input_raw/{segmento}/ y compara filas leidas en crudo
contra filas normalizadas.

Detecta perdidas inesperadas de datos en el pipeline ETL.

Uso:
  python src/normalizador/comparativo_registros.py
  python src/normalizador/comparativo_registros.py --segmento 2026
  python src/normalizador/comparativo_registros.py --segmento 2025 2026
"""

import argparse
import os
import sys
from datetime import datetime
from typing import Any

import openpyxl
import pandas as pd

INPUT_DIR = "fuentes/compensaciones"
OUTPUT_DIR = "output/output_ready_for_pq"
SEP = "=" * 68

SEGMENTOS_VALIDOS = [
    "2021_2022_Historico",
    "Modificaciones",
    "2025_Compensaciones",
    "2023_2025_Hist",
    "2025_Ajustes",
    "2025",
    "2026",
]


def parsear_argumentos() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Comparativo crudo -> ingesta pre-Power Query"
    )
    parser.add_argument(
        "--segmento",
        nargs="+",
        metavar="SEGMENTO",
        choices=SEGMENTOS_VALIDOS,
        help=("Segmento(s) a comparar. " "Si se omite, se comparan todos."),
    )
    return parser.parse_args()


def seleccionar_hoja(wb: openpyxl.Workbook) -> str:
    """
    Selecciona la hoja de datos del crudo con la misma
    prioridad que ExcelReader en reader.py.
    """
    sheet_map = {name.upper(): name for name in list(wb.sheetnames)}

    for name_upper in sheet_map:
        if name_upper.startswith("ANEXAR"):
            return sheet_map[name_upper]
    if "PRONTOPOSE_LIMPIA" in sheet_map:
        return sheet_map["PRONTOPOSE_LIMPIA"]
    if "BASE DE DATOS" in sheet_map:
        return sheet_map["BASE DE DATOS"]
    if "BASE" in sheet_map:
        return sheet_map["BASE"]
    if "INFORME MENSUAL" in sheet_map:
        return sheet_map["INFORME MENSUAL"]
    if "TABLA" in sheet_map:
        return sheet_map["TABLA"]
    return list(wb.sheetnames)[0]


def contar_filas_crudo(ruta: str) -> tuple[int, str]:
    """
    Cuenta filas de datos (sin cabecera) en el xlsx crudo.
    Retorna (filas, nombre_hoja).
    """
    wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
    hoja_nombre = seleccionar_hoja(wb)
    ws = wb[hoja_nombre]
    filas = max(0, (ws.max_row or 1) - 1)
    wb.close()
    return filas, hoja_nombre


def contar_filas_ingesta(ruta: str) -> int:
    """Cuenta filas en el xlsx normalizado (output_ready_for_pq)."""
    df = pd.read_excel(ruta, engine="openpyxl")
    return len(df)


def sumar_importe_crudo(ruta: str, hoja: str) -> float | None:
    """
    Suma el importe en el xlsx crudo.
    Usa IMPORTE2 si existe; si no, IMPORTE.
    Retorna None si ninguna columna está disponible.
    """
    df = pd.read_excel(ruta, sheet_name=hoja, engine="openpyxl")
    cols_up = [c.upper() for c in df.columns]
    if "IMPORTE2" in cols_up:
        col = df.columns[cols_up.index("IMPORTE2")]
    elif "IMPORTE 2" in cols_up:
        col = df.columns[cols_up.index("IMPORTE 2")]
    elif "IMPORTE" in cols_up:
        col = df.columns[cols_up.index("IMPORTE")]
    else:
        return None
    return float(pd.to_numeric(df[col], errors="coerce").sum())


def sumar_importe_ingesta(ruta: str) -> float | None:
    """
    Suma la columna IMPORTE en el xlsx normalizado.
    Retorna None si la columna no existe.
    """
    df = pd.read_excel(ruta, engine="openpyxl")
    cols_up = [c.upper() for c in df.columns]
    if "IMPORTE" not in cols_up:
        return None
    col = df.columns[cols_up.index("IMPORTE")]
    return float(pd.to_numeric(df[col], errors="coerce").sum())


def contar_obras_ingesta(ruta: str) -> int | None:
    """
    Cuenta valores únicos de OBRA_PRONTO en el xlsx normalizado.
    Retorna None si la columna no existe.
    """
    df = pd.read_excel(ruta, engine="openpyxl")
    cols_up = [c.upper() for c in df.columns]
    if "OBRA_PRONTO" not in cols_up:
        return None
    col = df.columns[cols_up.index("OBRA_PRONTO")]
    return int(df[col].dropna().nunique())


def contar_cuentas_ingesta(ruta: str) -> int | None:
    """
    Cuenta valores únicos de CODIGO_CUENTA en el xlsx normalizado.
    Retorna None si la columna no existe.
    """
    df = pd.read_excel(ruta, engine="openpyxl")
    cols_up = [c.upper() for c in df.columns]
    if "CODIGO_CUENTA" not in cols_up:
        return None
    col = df.columns[cols_up.index("CODIGO_CUENTA")]
    return int(df[col].dropna().nunique())


def listar_fuentes_ingesta(ruta: str) -> set[str]:
    """
    Retorna el conjunto de valores únicos de FUENTE en el xlsx
    normalizado. Retorna set vacío si la columna no existe.
    """
    df = pd.read_excel(ruta, engine="openpyxl")
    cols_up = [c.upper() for c in df.columns]
    if "FUENTE" not in cols_up:
        return set()
    col = df.columns[cols_up.index("FUENTE")]
    return set(df[col].dropna().astype(str).unique())


def fmt_imp(valor: float, ancho: int = 18, signo: bool = False) -> str:
    """
    Formatea un importe con separadores ES:
    punto (.) para miles, coma (,) para decimales.
    """
    fmt = f"+{ancho},.0f" if signo else f"{ancho},.0f"
    us = format(valor, fmt)
    # . -> # (temp), , -> . (miles), # -> , (decimal)
    es = us.replace(".", "#").replace(",", ".").replace("#", ",")
    return es


def recolectar_pares(
    segmentos: list[str] | None,
) -> list[dict[str, Any]]:
    """
    Retorna lista de dicts con rutas crudo/ingesta por archivo.
    Solo incluye archivos presentes en output_ready_for_pq.
    """
    resultado: list[dict[str, Any]] = []
    dirs_seg = segmentos if segmentos else sorted(os.listdir(OUTPUT_DIR))

    for seg in dirs_seg:
        dir_out = os.path.join(OUTPUT_DIR, seg)
        dir_in = os.path.join(INPUT_DIR, seg)
        if not os.path.isdir(dir_out):
            continue
        for nombre in sorted(os.listdir(dir_out)):
            if not nombre.lower().endswith(".xlsx"):
                continue
            ruta_out = os.path.join(dir_out, nombre)
            ruta_in = os.path.join(dir_in, nombre)
            resultado.append(
                {
                    "segmento": seg,
                    "nombre": nombre,
                    "ruta_out": ruta_out,
                    "ruta_in": ruta_in if os.path.exists(ruta_in) else None,
                }
            )
    return resultado


def main() -> None:
    args = parsear_argumentos()
    segmentos: list[str] | None = args.segmento

    print()
    print(SEP)
    print("  COMPARATIVO CRUDO -> INGESTA")
    scope = ", ".join(segmentos) if segmentos else "todos los segmentos"
    print(
        f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}" f"  |  scope: {scope}"
    )
    print(SEP)

    pares = recolectar_pares(segmentos)
    if not pares:
        print("  No se encontraron archivos para comparar.")
        sys.exit(0)

    bloqueantes: list[str] = []
    avisos: list[str] = []
    seg_actual = ""

    for par in pares:
        if par["segmento"] != seg_actual:
            seg_actual = par["segmento"]
            print(f"\n  {seg_actual}")
            print(
                f"     {'Archivo':<40} {'Crudo':>8}"
                f" {'Ingesta':>8}  {'Delta':>7}  Estado"
            )
            print("     " + "-" * 72)

        nombre: str = par["nombre"]
        ruta_out: str = par["ruta_out"]
        ruta_in: str | None = par["ruta_in"]
        clave = f"{seg_actual}/{nombre}"

        if ruta_in is None:
            print(
                f"     {nombre:<40} {'---':>8}"
                f" {'---':>8}  {'---':>7}"
                f"  sin crudo"
            )
            avisos.append(clave)
            continue

        try:
            filas_in, hoja = contar_filas_crudo(ruta_in)
            filas_out = contar_filas_ingesta(ruta_out)
            delta = filas_out - filas_in

            if filas_in > 0 and filas_out == 0:
                estado = "PERDIDA TOTAL"
                bloqueantes.append(clave)
            elif delta < 0:
                pct = abs(delta) / filas_in * 100
                estado = f"BAJA {pct:.1f}%"
                avisos.append(clave)
            elif delta == 0:
                estado = "OK"
            else:
                estado = f"INFO +{delta:,}"

            signo = "+" if delta >= 0 else ""
            print(
                f"     {nombre:<40} {filas_in:>8,}"
                f" {filas_out:>8,}  {signo}{delta:>6,}"
                f"  {estado}"
            )
            imp_in = sumar_importe_crudo(ruta_in, hoja)
            imp_out = sumar_importe_ingesta(ruta_out)
            if imp_in is not None and imp_out is not None:
                d_imp = imp_out - imp_in
                e_imp = "OK" if abs(d_imp) < 0.01 else "DIFERENCIA"
                s_in = fmt_imp(imp_in, 18)
                s_out = fmt_imp(imp_out, 18)
                s_d = fmt_imp(d_imp, 14, signo=True)
                print(
                    f"     {'  Σ IMPORTE':<40}"
                    f" {s_in}"
                    f" {s_out}"
                    f"  {s_d}"
                    f"  {e_imp}"
                )

            # ── Controles adicionales (ingesta) ──────────────────────
            obras_u = contar_obras_ingesta(ruta_out)
            if obras_u is not None:
                e_obras = "OK" if obras_u > 0 else "SIN OBRAS"
                if obras_u == 0:
                    bloqueantes.append(f"{clave} [0 obras]")  # noqa: E501
                print(
                    f"     {'  Obras únicas':<40}"
                    f" {'':>18} {obras_u:>8,}"
                    f"  {'':>14}  {e_obras}"
                )

            cuentas_u = contar_cuentas_ingesta(ruta_out)
            if cuentas_u is not None:
                e_cc = "OK" if cuentas_u > 0 else "SIN CUENTAS"
                if cuentas_u == 0:
                    bloqueantes.append(f"{clave} [0 cuentas]")
                print(
                    f"     {'  Cuentas únicas':<40}"
                    f" {'':>18} {cuentas_u:>8,}"
                    f"  {'':>14}  {e_cc}"
                )

            fuentes = listar_fuentes_ingesta(ruta_out)
            if fuentes:
                fuentes_str = ", ".join(sorted(fuentes))
                print(f"     {'  Fuentes':<40}" f" {fuentes_str}")
        except Exception as e:
            print(
                f"     {nombre:<40} {'---':>8}"
                f" {'---':>8}  {'---':>7}"
                f"  Error: {e}"
            )
            bloqueantes.append(clave)

    print()
    print(SEP)
    if bloqueantes:
        print("  COMPARATIVO BLOQUEANTE - perdida total en:")
        for b in bloqueantes:
            print(f"     - {b}")
        print(
            "  Pipeline detenido. " "Revisar normalizador antes de continuar."
        )
        print(SEP)
        print()
        sys.exit(1)
    if avisos:
        print("  Caidas parciales (filtros de normalizacion):")
        for a in avisos:
            print(f"     - {a}")
        print("  Revisar si los registros descartados son esperados.")
    if not bloqueantes and not avisos:
        print("  Comparativo OK - sin perdidas detectadas.")
    print(SEP)
    print()


if __name__ == "__main__":
    main()
